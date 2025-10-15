"""
confluence_client.py
Provides a ConfluenceClient class to connect to Atlassian Confluence and fetch page content and attachments by page title.
"""

from atlassian import Confluence
from dotenv import load_dotenv
import html2text
import os
from docx import Document
from PyPDF2 import PdfReader
from openpyxl import load_workbook
import io
import json
import requests
import logging
import sys
from typing import Dict, Any
import argparse

SUPPORTED_TEXT_TYPES = [
    "text/",
    "application/json",
    "application/pdf",
    "application/vnd.openxmlformats-officedocument",
    "application/msword",
    "application/vnd.ms-excel",
    "application/vnd.ms-powerpoint",
]
SUPPORTED_TEXT_EXTS = (".txt", ".json", ".pdf", ".docx", ".xlsx")
DRAWIO_MIME = "application/vnd.jgraph.mxfile"
DRAWIO_EXTS = (".drawio",)
IMAGE_EXTS = (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".svg", ".webp")

# Logging configuration
logging.basicConfig(
    level=logging.INFO,
    format="❌%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()],
)


class ConfluenceClient:
    """
    ConfluenceClient provides methods to connect to Confluence and fetch page content and attachments.
    """

    def __init__(self):
        """
        Initialize Confluence client using environment variables.
        """
        load_dotenv()
        self.url = os.getenv("CONFLUENCE_URL")
        self.api_token = os.getenv("CONFLUENCE_ACCESS_TOKEN")
        self.recursive_depth = int(os.getenv("CONFLUENCE_RECURSIVE_DEPTH", "1"))
        if not self.url or not self.api_token:
            raise ValueError(
                "❌ Missing CONFLUENCE_URL or CONFLUENCE_ACCESS_TOKEN in environment variables."
            )
        self.confluence = Confluence(url=self.url, token=self.api_token)

    def get_confluence_details(self, page_title: str) -> str:
        """
        Fetches the content and parsed attachments of a Confluence page by its title.
        :param page_title: The title of the Confluence page
        :return: Markdown string with page content and attachment details
        """
        page = self._get_page_by_title(page_title)
        markdown = html2text.html2text(page["body"]["storage"]["value"])
        attachments = self._get_attachments_and_parse(page)
        if attachments:
            markdown += f"\n\n# Attachments (Total: {len(attachments)})\n"
            for fname, content in attachments.items():
                markdown += f"\n## {fname}\n\n```\n{content}\n```\n"
        return markdown

    def get_confluence_details_recursive(
        self, page_title: str, depth: int = None, current_level: int = 0
    ) -> str:
        if depth is None:
            depth = self.recursive_depth
        page = self._get_page_by_title(page_title)
        markdown = f"{'#' * (current_level + 1)} {page_title}\n"
        markdown += html2text.html2text(page["body"]["storage"]["value"])
        attachments = self._get_attachments_and_parse(page)
        if attachments:
            markdown += f"\n\n## Attachments (Total: {len(attachments)})\n"
            for fname, content in attachments.items():
                markdown += f"\n### {fname}\n\n```\n{content}\n```\n"
        if current_level < depth - 1:
            child_pages = self._get_child_pages(page["id"])
            for child in child_pages:
                child_title = child["title"]
                markdown += "\n" + self.get_confluence_details_recursive(
                    child_title, depth, current_level + 1
                )
        return markdown

    def _get_child_pages(self, page_id: str) -> list:
        children = self.confluence.get_child_content(page_id, type="page").get(
            "results", []
        )
        return children

    def _get_page_by_title(self, page_title: str) -> Dict[str, Any]:
        """
        Fetch a Confluence page by its title.
        :param page_title: The title of the Confluence page
        :return: Page dictionary
        """
        page = self.confluence.get_page_by_title(
            space=None, title=page_title, expand="body.storage"
        )
        if not page or "body" not in page or "storage" not in page["body"]:
            raise ValueError(
                f"❌ Could not fetch content for page title '{page_title}'"
            )
        return page

    def _get_attachments_and_parse(self, page: Dict[str, Any]) -> Dict[str, str]:
        """
        Download and parse all attachments for a given Confluence page.
        :param page: Page dictionary
        :return: Dict mapping filename to extracted or parsed content
        """
        if not page or "id" not in page:
            return {}
        page_id = page["id"]
        attachments = self.confluence.get_attachments_from_content(page_id).get(
            "results", []
        )
        parsed = {}
        for att in attachments:
            filename = att["title"]
            # Skip download for images/binaries by extension and temp draw.io files
            if filename.lower().endswith(IMAGE_EXTS) or filename.lower().endswith(
                ".tmp"
            ):
                logging.info(f"Skipping image or temp file: {filename}")
                parsed[filename] = "[File not displayed: image or temp file omitted]"
                continue
            download_link = att["_links"]["download"]
            url = self.url.rstrip("/") + download_link
            headers = {"Authorization": f"Bearer {self.api_token}"}
            try:
                response = requests.get(url, headers=headers)
                file_content = response.content
                content_type = response.headers.get("Content-Type", "").lower().strip()
            except Exception as e:
                logging.warning(f"❌ Failed to download {filename}: {e}")
                parsed[filename] = f"❌ [Download error: {e}]"
                continue
            # Now check content type and process
            if self._is_drawio(filename, content_type):
                xml_str = file_content.decode("utf-8", errors="replace")
                parsed[filename] = (
                    "Refer the attached the Draw.io XML content : \n " + xml_str
                )
            elif self._is_supported_text(filename, content_type):
                parsed[filename] = self._extract_text_from_attachment(
                    filename, file_content
                )
            else:
                logging.warning(
                    f"Skipping unsupported file after download: {filename} ({content_type or 'unknown'})"
                )
                parsed[filename] = (
                    f'[File type not parsed: {content_type or "unknown"}]'
                )
        return parsed

    def _is_drawio(self, filename: str, content_type: str) -> bool:
        """
        Check if a file is a draw.io diagram based on filename or content type.
        :param filename: Filename
        :param content_type: Content type string
        :return: True if draw.io, else False
        """
        return (
            content_type.startswith(DRAWIO_MIME)
            or filename.endswith(DRAWIO_EXTS)
            or (filename.endswith(".xml") and "drawio" in filename.lower())
        )

    def _is_supported_text(self, filename: str, content_type: str) -> bool:
        """
        Check if a file is a supported text type.
        :param filename: Filename
        :param content_type: Content type string
        :return: True if supported text, else False
        """
        return any(
            content_type.startswith(ct) for ct in SUPPORTED_TEXT_TYPES
        ) or filename.lower().endswith(SUPPORTED_TEXT_EXTS)

    def _extract_text_from_attachment(self, filename: str, raw_content: bytes) -> str:
        """
        Extract text from an attachment based on file extension.
        :param filename: Filename
        :param raw_content: Raw file bytes
        :return: Extracted text or a message for unsupported types
        """
        filename = filename.lower()
        if filename.endswith(
            (".png", ".jpg", ".jpeg", ".gif", ".bmp", ".svg", ".webp")
        ):
            return "[Image file: not displayed, binary content omitted]"
        elif filename.endswith(".xlsx"):
            wb = load_workbook(io.BytesIO(raw_content), data_only=True)
            text = []
            for sheet in wb.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    text.append(
                        "\t".join(
                            [str(cell) if cell is not None else "" for cell in row]
                        )
                    )
            return "\n".join(text)
        elif filename.endswith(".docx"):
            doc = Document(io.BytesIO(raw_content))
            return "\n".join([para.text for para in doc.paragraphs])
        elif filename.endswith(".pdf"):
            reader = PdfReader(io.BytesIO(raw_content))
            text = []
            for page in reader.pages:
                text.append(page.extract_text() or "")
            return "\n".join(text)
        elif filename.endswith(".json"):
            try:
                data = json.loads(raw_content.decode(errors="replace"))
                return json.dumps(data, indent=2)
            except Exception:
                return raw_content.decode(errors="replace")
        elif filename.endswith(".txt"):
            return raw_content.decode(errors="replace")
        else:
            return raw_content.decode(errors="replace")[:100000]


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fetch Confluence details <page_title>."
    )
    parser.add_argument(
        "page_title",
        nargs="?",
        default="Phase 1: Rebuild of the IFrame and Address (MAP)",
        help="Confluence Page Title",
    )
    parser.add_argument(
        "--recursive",
        action="store_true",
        help="Recursively fetch child pages up to configured depth",
    )
    args = parser.parse_args()
    page_title = (
        args.page_title
        if getattr(args, "page_title", None)
        else "Phase 1: Rebuild of the IFrame and Address (MAP)"
    )
    client = ConfluenceClient()
    try:
        if args.recursive:
            content = client.get_confluence_details_recursive(page_title)
        else:
            content = client.get_confluence_details(page_title)
        if content:
            print(
                f"✅ Confluence page '{page_title}' retrieved successfully:\n{content}"
            )
        else:
            print(f"❌ Failed to retrieve page {page_title}")
            sys.exit(1)
    except Exception as e:
        print(f"❌ Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
