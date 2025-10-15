import os
import sys
import io
import json
import logging
import argparse
from typing import Optional, Dict, Any, List

from jira import JIRA
from dotenv import load_dotenv
from openpyxl import load_workbook
from docx import Document
from PyPDF2 import PdfReader
import magic

from PIL import Image
import pytesseract
import tempfile
from moviepy import VideoFileClip
import speech_recognition as sr

from transformers import BlipProcessor, BlipForConditionalGeneration
import torch

# Constants
ATTACHMENT_CHAR_LIMIT = 100000
ACCEPTANCE_CRITERIA_FIELD = os.getenv(
    "jira_acceptance_criteria_field", "customfield_10700"
)

# Logging configuration
logging.basicConfig(
    level=logging.INFO,
    format="❌%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()],
)


class JiraClient:
    """
    JiraClient provides methods to connect to Jira, fetch issue details, and extract text/data from attachments.
    """

    def __init__(
        self, server: Optional[str] = None, access_token: Optional[str] = None
    ):
        """
        Initialize Jira client using environment variables or provided parameters.
        :param server: Jira server URL (optional, defaults to env)
        :param access_token: Jira access token (optional, defaults to env)
        """
        load_dotenv()
        self.image_transformation = (
            os.getenv("JIRA_IMAGE_TRANSFORMATION", "false").lower() == "true"
        )
        self.video_transformation = (
            os.getenv("JIRA_VIDEO_TRANSFORMATION", "false").lower() == "true"
        )

        self.server = server or os.getenv("JIRA_SERVER_URL")
        self.access_token = access_token or os.getenv("JIRA_ACCESS_TOKEN")
        if not self.server or not self.access_token:
            logging.critical(
                "Jira server URL and access token must be provided via env or parameters."
            )
            raise ValueError(
                "Jira server URL and access token must be provided via env or parameters."
            )
        try:
            self.jira = JIRA(server=self.server, token_auth=self.access_token)
        except Exception as e:
            logging.critical(f"❌ Failed to initialize JIRA client: {e}")
            raise

    def get_issue_details(self, jira_id: str) -> str:
        """
        Get issue details for the given jira id from Jira.
        :param jira_id: The Jira issue ID (e.g., MSDP-218)
        :return: Formatted string with issue details and extracted attachment content
        """
        try:
            issue = self.jira.issue(jira_id)
        except Exception as e:
            logging.error(f"Error fetching issue {jira_id}: {e}")
            return f"❌ Unable to fetch issue details: {e}"
        if not issue:
            return f"❌ Unable to fetch issue details."
        return self.format_issue_details(issue)

    def format_issue_details(self, issue: Any) -> str:
        """
        Format and return details of a Jira issue, including extracted attachment content.
        :param issue: Jira issue object
        :return: Formatted string with issue details
        """
        details: List[str] = []
        try:
            project = getattr(issue.fields.project, "name", "N/A")
            issue_type = getattr(issue.fields.issuetype, "name", "N/A")
            summary = getattr(issue.fields, "summary", "N/A")
            details.append(f"📦 Project: {project}")
            details.append(f"📦 Issue Type: {issue_type}")
            details.append(f"📦 Summary: {summary}")
            description = getattr(issue.fields, "description", None)
            if description:
                details.append(f"📦 Details: {description}")
            acceptance_criteria = getattr(issue.fields, ACCEPTANCE_CRITERIA_FIELD, None)
            if acceptance_criteria:
                details.append(f"📦 Acceptance Criteria: {acceptance_criteria}")
            attachments = getattr(issue.fields, "attachment", None)
            if attachments:
                details.append(f"📦 Attachments: (Total: {len(attachments)})")
                extracted_map = self.get_all_attachments_extracted_text(attachments)
                for idx, attachment in enumerate(attachments, 1):
                    details.append(
                        f"📦 {idx}. Attachment File: {attachment.filename} (Uploaded by: {getattr(attachment.author, 'displayName', 'N/A')})"
                    )
                    extracted = extracted_map.get(attachment.filename)
                    if extracted:
                        details.append(
                            f"📦   Extracted content (first {ATTACHMENT_CHAR_LIMIT} chars):\n{extracted[:ATTACHMENT_CHAR_LIMIT]}"
                        )
        except Exception as e:
            logging.error(f"❌ Error formatting issue details: {e}")
            details.append(f"❌ [Error formatting issue details: {e}]")
        return "\n".join(details)

    def get_all_attachments_extracted_text(
        self, attachments: List[Any]
    ) -> Dict[str, str]:
        """
        Return a dict of filename to extracted text/data for all attachments.
        :param attachments: List of Jira attachment objects
        :return: Dict mapping filename to extracted text
        """
        extracted_dict: Dict[str, str] = {}
        for attachment in attachments:
            try:
                extracted = self.extract_text_from_attachment(attachment)
                extracted_dict[attachment.filename] = extracted
            except Exception as e:
                logging.warning(
                    f"❌ Error extracting content from {attachment.filename}: {e}"
                )
                extracted_dict[attachment.filename] = (
                    f"❌ Error extracting content: {e}"
                )
        return extracted_dict

    def extract_text_from_attachment(self, attachment: Any) -> Optional[str]:
        """
        Extract text/data from an attachment based on its file type.
        Supports: .xlsx, .docx, .pdf, .json, .txt, and others (as text).
        :param attachment: Jira attachment object
        :return: Extracted text or None
        """
        filename = attachment.filename.lower()
        raw_content = self.get_attachment_content(attachment)
        if raw_content is None:
            return None
        content_type = magic.from_buffer(raw_content, mime=True)
        handler_map = {
            ".xlsx": self._extract_xlsx,
            ".docx": self._extract_docx,
            ".pdf": self._extract_pdf,
            ".json": self._extract_json,
            ".txt": self._extract_txt,
            ".jpg": self._extract_jpeg,
            ".jpeg": self._extract_jpeg,
            ".mp4": self._extract_mp4,
        }
        for ext, handler in handler_map.items():
            if filename.endswith(ext):
                return handler(raw_content)
            if filename.endswith((".jpg", ".jpeg")):
                if self.image_transformation:
                    return self._extract_jpeg(raw_content)
                else:
                    return "🔍[Image transformation disabled by config]"
            if filename.endswith(".mp4"):
                if self.video_transformation:
                    return self._extract_mp4(raw_content)
                else:
                    return "🔍[Video transformation disabled by config]"

        # Fallback for other text-like files
        if content_type.startswith("text/") or content_type == "application/json":
            return raw_content.decode(errors="replace")[:ATTACHMENT_CHAR_LIMIT]
        return f'❌ [File type not parsed: {content_type or "unknown"}]'

    def get_attachment_content(self, attachment: Any) -> Optional[bytes]:
        """
        Download and return the raw content of a Jira attachment object using the SDK.
        Handles both file-like and iterable cases for compatibility.
        :param attachment: Jira attachment object
        :return: Raw bytes of the attachment or None
        """
        try:
            file_obj = attachment.get()
            if hasattr(file_obj, "read"):
                return file_obj.read()
            return bytes(file_obj)
        except Exception as e:
            logging.error(f"❌ Error downloading attachment: {e}")
            return None

    # --- File type handlers ---
    def _extract_xlsx(self, raw_content: bytes) -> str:
        """Extract text from an Excel (.xlsx) file."""
        wb = load_workbook(io.BytesIO(raw_content), data_only=True)
        text: List[str] = []
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(values_only=True):
                text.append(
                    "\t".join([str(cell) if cell is not None else "" for cell in row])
                )
        return "\n".join(text)

    def _extract_docx(self, raw_content: bytes) -> str:
        """Extract text from a Word (.docx) file."""
        doc = Document(io.BytesIO(raw_content))
        return "\n".join([para.text for para in doc.paragraphs])

    def _extract_pdf(self, raw_content: bytes) -> str:
        """Extract text from a PDF file."""
        reader = PdfReader(io.BytesIO(raw_content))
        text: List[str] = []
        for page in reader.pages:
            text.append(page.extract_text() or "")
        return "\n".join(text)

    def _extract_json(self, raw_content: bytes) -> str:
        """Extract and pretty-print JSON content."""
        try:
            data = json.loads(raw_content.decode(errors="replace"))
            return json.dumps(data, indent=2)
        except Exception:
            return raw_content.decode(errors="replace")

    def _extract_txt(self, raw_content: bytes) -> str:
        """Extract text from a plain text file."""
        return raw_content.decode(errors="replace")

    # def _extract_jpeg(self, raw_content: bytes) -> str:
    #     """Extract text from a JPEG image using OCR."""
    #     with tempfile.NamedTemporaryFile(suffix=".jpg") as temp_img:
    #         temp_img.write(raw_content)
    #         temp_img.flush()
    #         text = pytesseract.image_to_string(Image.open(temp_img.name))
    #     return text.strip() or "[No text detected in image]"

    def _extract_jpeg(self, raw_content: bytes) -> str:
        """Extract text from a JPEG image using OCR, or summarize with BLIP if no text found."""
        try:
            with tempfile.NamedTemporaryFile(suffix=".jpg") as temp_img:
                temp_img.write(raw_content)
                temp_img.flush()
                text = pytesseract.image_to_string(Image.open(temp_img.name)).strip()
                if text:
                    return text
                # Summarize image using BLIP
                try:
                    from transformers import BlipProcessor, BlipForConditionalGeneration
                    import torch

                    processor = BlipProcessor.from_pretrained(
                        "Salesforce/blip-image-captioning-base"
                    )
                    model = BlipForConditionalGeneration.from_pretrained(
                        "Salesforce/blip-image-captioning-base"
                    )
                    image = Image.open(temp_img.name).convert("RGB")
                    inputs = processor(image, return_tensors="pt")
                    with torch.no_grad():
                        out = model.generate(**inputs)
                    caption = processor.decode(out[0], skip_special_tokens=True)
                    return f"📦 [Image summary]: {caption}"
                except Exception as e:
                    return f"❌ [Image summary failed: {e}]"
        except Exception as e:
            return f"❌ [JPEG extraction failed: {e}]"

    def _extract_mp4(self, raw_content: bytes) -> str:
        """Extract transcript from an MP4 video using speech recognition."""
        with tempfile.NamedTemporaryFile(suffix=".mp4") as temp_vid:
            temp_vid.write(raw_content)
            temp_vid.flush()
            clip = VideoFileClip(temp_vid.name)
            if clip.audio is None or self.video_transformation:
                # Summarize video using BLIP
                return self._summarize_mp4_video(raw_content)
            # Otherwise, extract transcript as before
            audio_path = temp_vid.name + ".wav"
            clip.audio.write_audiofile(audio_path)
            recognizer = sr.Recognizer()
            with sr.AudioFile(audio_path) as source:
                audio = recognizer.record(source)
            try:
                transcript = recognizer.recognize_google(audio)
            except Exception as e:
                transcript = f"❌ [Speech recognition failed: {e}]"
        return transcript

    @staticmethod
    def _summarize_mp4_video(raw_content: bytes, num_frames: int = 5) -> str:
        """Summarize an MP4 video by generating captions for key frames."""
        with tempfile.NamedTemporaryFile(suffix=".mp4") as temp_vid:
            temp_vid.write(raw_content)
            temp_vid.flush()
            clip = VideoFileClip(temp_vid.name)
            duration = clip.duration
            times = [duration * i / (num_frames + 1) for i in range(1, num_frames + 1)]
            processor = BlipProcessor.from_pretrained(
                "Salesforce/blip-image-captioning-base"
            )
            model = BlipForConditionalGeneration.from_pretrained(
                "Salesforce/blip-image-captioning-base"
            )
            captions = []
            for t in times:
                frame = clip.get_frame(t)
                image = Image.fromarray(frame).convert("RGB")
                inputs = processor(image, return_tensors="pt")
                with torch.no_grad():
                    out = model.generate(**inputs)
                caption = processor.decode(out[0], skip_special_tokens=True)
                captions.append(f"📦 Frame at {t:.1f}s: {caption}")
            return "\n".join(captions)


def main() -> None:
    """
    Main entry point for running the script from the command line.
    Usage: python jira_client.py <jira_id>
    """
    parser = argparse.ArgumentParser(description="Fetch Jira issue details.")
    parser.add_argument(
        "jira_id", nargs="?", default="CLOV-974", help="Jira issue ID (e.g., MSDP-218)"
    )

    args = parser.parse_args()
    jira_id = args.jira_id if getattr(args, "jira_id", None) else "CLOV-974"

    jira_client = JiraClient()
    try:
        issue_details = jira_client.get_issue_details(jira_id)
        if issue_details:
            print(f"✅ Issue {jira_id} retrieved successfully:\n{issue_details}")
        else:
            print(f"❌ Failed to retrieve issue {jira_id}")
            sys.exit(1)
    except Exception as e:
        logging.error(f"❌ Fatal error: {e}")
        print(f"❌ Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
